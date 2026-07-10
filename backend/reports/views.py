from datetime import timedelta, datetime
from decimal import Decimal

from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.cache import never_cache
from django.utils import timezone
from django.db.models import Sum

from django.core.paginator import Paginator
from django.conf import settings

@never_cache
@staff_member_required(login_url='admin_login')
def sales_report(request):
    from orders.models import Order

    today = timezone.now().date()
    period = request.GET.get('period', 'daily')
    if period not in ('daily', 'weekly', 'yearly', 'custom'):
        period = 'daily'

    custom_start = request.GET.get('start_date', '').strip()
    custom_end   = request.GET.get('end_date', '').strip()

    if period == 'daily':
        start_date = today
        end_date   = today
    elif period == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        end_date   = start_date + timedelta(days=6)
    elif period == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date   = today.replace(month=12, day=31)
    else:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            start_date = today
        try:
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            end_date = today
        if end_date < start_date:
            start_date, end_date = end_date, start_date

    valid_statuses = ['pending', 'shipped', 'out_for_delivery', 'delivered']

    orders_in_range = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status__in=valid_statuses,
    ).select_related('user').order_by('-created_at')

    sales_count = orders_in_range.count()
    gross_amount = orders_in_range.aggregate(r=Sum('subtotal'))['r'] or Decimal('0.00')
    coupon_discount_total = orders_in_range.aggregate(r=Sum('coupon_discount'))['r'] or Decimal('0.00')
    total_discount_amount = orders_in_range.aggregate(r=Sum('discount_amount'))['r'] or Decimal('0.00')
    offer_discount_total = total_discount_amount - coupon_discount_total
    net_revenue = orders_in_range.aggregate(r=Sum('total_amount'))['r'] or Decimal('0.00')

    # Pagination 
    print("Total orders:", orders_in_range.count())
    paginator = Paginator(orders_in_range, getattr(settings, 'REPORTS_PER_PAGE', 5))
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'sales_count': sales_count,
        'gross_amount': gross_amount,
        'offer_discount_total': offer_discount_total,
        'coupon_discount_total': coupon_discount_total,
        'total_discount_amount': total_discount_amount,
        'net_revenue': net_revenue,
        'orders': page_obj,
        'page_obj': page_obj,
        'today': today,
    }
    return render(request, 'reports/sales_report.html', context)

from django.http import HttpResponse


def _get_report_data(request):
    
    from orders.models import Order

    today = timezone.now().date()
    period = request.GET.get('period', 'daily')
    if period not in ('daily', 'weekly', 'yearly', 'custom'):
        period = 'daily'

    custom_start = request.GET.get('start_date', '').strip()
    custom_end   = request.GET.get('end_date', '').strip()

    if period == 'daily':
        start_date = today
        end_date   = today
    elif period == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        end_date   = start_date + timedelta(days=6)
    elif period == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date   = today.replace(month=12, day=31)
    else:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            start_date = today
        try:
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            end_date = today
        if end_date < start_date:
            start_date, end_date = end_date, start_date

    valid_statuses = ['pending', 'shipped', 'out_for_delivery', 'delivered']

    orders_in_range = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        status__in=valid_statuses,
    ).select_related('user').order_by('-created_at')

    sales_count = orders_in_range.count()
    gross_amount = orders_in_range.aggregate(r=Sum('subtotal'))['r'] or Decimal('0.00')
    coupon_discount_total = orders_in_range.aggregate(r=Sum('coupon_discount'))['r'] or Decimal('0.00')
    total_discount_amount = orders_in_range.aggregate(r=Sum('discount_amount'))['r'] or Decimal('0.00')
    offer_discount_total = total_discount_amount - coupon_discount_total
    net_revenue = orders_in_range.aggregate(r=Sum('total_amount'))['r'] or Decimal('0.00')

    return {
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'sales_count': sales_count,
        'gross_amount': gross_amount,
        'offer_discount_total': offer_discount_total,
        'coupon_discount_total': coupon_discount_total,
        'net_revenue': net_revenue,
        'orders': orders_in_range,
        'total_discount_amount': total_discount_amount,
    }


@never_cache
@staff_member_required(login_url='admin_login')
def sales_report_pdf(request):
    from weasyprint import HTML
    from django.template.loader import render_to_string

    data = _get_report_data(request)
    data['generated_by'] = request.user.email
    data['generated_at'] = timezone.now().strftime('%d %b %Y, %I:%M %p')

    html_string = render_to_string('reports/sales_report_pdf.html', data)
    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"sales-report-{data['start_date']}-to-{data['end_date']}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@never_cache
@staff_member_required(login_url='admin_login')
def sales_report_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    data = _get_report_data(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=16, color="111827")
    subtitle_font = Font(size=10, color="6b7280")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb'),
    )
    currency_format = '"₹"#,##0.00'

    
    ws.merge_cells('A1:F1')
    ws['A1'] = "TechGear"
    ws['A1'].font = title_font

    ws.merge_cells('A2:F2')
    ws['A2'] = "Sales Report — Internal Use"
    ws['A2'].font = subtitle_font

    ws.merge_cells('A3:F3')
    ws['A3'] = (
        f"Period: {data['start_date'].strftime('%d %b %Y')} to {data['end_date'].strftime('%d %b %Y')} "
        f"| Report type: {data['period'].title()}"
    )
    ws['A3'].font = subtitle_font

    
    ws.append([])
    summary_header_row = 5
    ws.append(['Sales Count', 'Gross Amount', 'Offer Discount', 'Coupon Discount', 'Net Revenue'])
    for cell in ws[summary_header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws.append([
        data['sales_count'],
        float(data['gross_amount']),
        float(data['offer_discount_total']),
        float(data['coupon_discount_total']),
        float(data['net_revenue']),
    ])
    summary_data_row = summary_header_row + 1
    for col in range(2, 6):
        cell = ws.cell(row=summary_data_row, column=col)
        cell.number_format = currency_format
        cell.border = thin_border
    ws.cell(row=summary_data_row, column=1).border = thin_border
    ws.cell(row=summary_data_row, column=1).alignment = Alignment(horizontal='center')

    
    ws.append([])
    ws.append(['Order #', 'Customer', 'Date', 'Status', 'Discount', 'Amount'])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for order in data['orders']:
        ws.append([
            order.order_number,
            order.user.email if order.user else 'Deleted User',
            order.created_at.strftime('%d %b %Y'),
            order.get_status_display(),
            float(order.discount_amount),
            float(order.total_amount),
        ])
        row = ws.max_row
        ws.cell(row=row, column=5).number_format = currency_format
        ws.cell(row=row, column=6).number_format = currency_format
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border

    
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=4, value="Total").font = bold_font
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
    discount_sum_cell = ws.cell(row=total_row, column=5, value=f"=SUM(E{header_row+1}:E{total_row-1})")
    discount_sum_cell.font = bold_font
    discount_sum_cell.number_format = currency_format
    amount_sum_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{total_row-1})")
    amount_sum_cell.font = bold_font
    amount_sum_cell.number_format = currency_format

    
    ws.auto_filter.ref = f"A{header_row}:F{total_row-1}"
    ws.freeze_panes = f"A{header_row + 1}"

    
    widths = {1: 16, 2: 30, 3: 14, 4: 18, 5: 14, 6: 14}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    
    footer_row = total_row + 2
    ws.merge_cells(f'A{footer_row}:F{footer_row}')
    footer_cell = ws.cell(
        row=footer_row, column=1,
        value=f"Generated by {request.user.email} on {timezone.now().strftime('%d %b %Y, %I:%M %p')}"
    )
    footer_cell.font = Font(size=8, color="9ca3af", italic=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"sales-report-{data['start_date']}-to-{data['end_date']}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response